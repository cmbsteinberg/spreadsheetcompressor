import logging
import openpyxl
from pathlib import Path
from typing import Dict, List, Optional, Union, Tuple
import re
import csv
import requests
from datetime import datetime
from collections import defaultdict
import pyexcel_ods
from io import BytesIO, StringIO
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from packaging.version import Version


# Define a custom exception for file parsing errors
class SpreadsheetParsingError(Exception):
    """Custom exception raised when spreadsheet parsing fails."""

    pass


class SpreadsheetCompressor:
    """
    A class to parse and compress spreadsheet data for LLM consumption.

    Supports Excel (.xlsx, .xlsm, .xltx, .xltm), ODS (.ods), and CSV (.csv) files.
    Compresses cell references into ranges where possible and groups data by type.
    """

    SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm", ".xltx", ".xltm", ".ods", ".csv")

    def __init__(
        self,
        log_level: int = logging.INFO,
        custom_patterns: Optional[Dict[str, str]] = None,
        custom_date_patterns: Optional[List[str]] = None,
        custom_time_patterns: Optional[List[str]] = None,
    ):
        """
        Initialize the SpreadsheetCompressor with logging and custom pattern configurations.

        Args:
            log_level: The logging level to use (default: logging.INFO).
            custom_patterns: A dictionary of custom regex patterns to recognize data types.
                             Keys are pattern names (e.g., "CUSTOM_ID"), and values are the regex strings.
            custom_date_patterns: A list of custom date format strings to try during data type recognition.
            custom_time_patterns: A list of custom time format strings to try during data type recognition.
        """
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(log_level)
        if not self.logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter(
                "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
            )
            handler.setFormatter(formatter)
            self.logger.addHandler(handler)

        # Compile default regex patterns
        default_patterns = {
            "url": r"^(https?:\/\/|www\.|ftp:\/\/|file:\/\/)[a-zA-Z0-9\-\.]+\.[a-zA-Z]{2,}(:[0-9]+)?(\/\S*)?$",
            "year": r"^(1|2)\d{3}$",
            "integer": r"^-?\d+$",
            "percentage": r"^-?\d+\.?\d*%$",
            "scientific": r"^-?\d+\.?\d*[eE][+-]?\d+$",
            "float": r"^-?\d*\.\d+$",
            "currency": r"^[$€£¥]\s*-?\d+\.?\d*$|^-?\d+\.?\d*\s*[$€£¥]$",
            "email": r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$",
        }
        self._patterns = {**default_patterns, **(custom_patterns or {})}

        # Define default date and time patterns
        self._date_patterns = custom_date_patterns or [
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%m-%d-%Y",
            "%Y/%m/%d",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%d.%m.%Y",
            "%Y.%m.%d",
            "%b %d, %Y",
            "%B %d, %Y",
            "%d %b %Y",
            "%d %B %Y",
        ]
        self._time_patterns = custom_time_patterns or [
            "%H:%M",
            "%H:%M:%S",
            "%I:%M %p",
            "%I:%M:%S %p",
        ]

        self._excel_min_version = Version(
            "1.4.0"
        )  # Minimum version for openpyxl features

    def recognize_data_type(self, value: Optional[Union[str, int, float]]) -> str:
        """
        Recognize the data type of a cell value.

        Args:
            value: The cell value.

        Returns:
            str: The recognized data type (e.g., "[INTEGER]", "[DATE]", "Others").
        """
        if value is None or (isinstance(value, str) and not value.strip()):
            return "Empty"

        value_str = str(value).strip()

        # Check against compiled patterns
        for pattern_name, pattern in self._patterns.items():
            if re.match(pattern, value_str):
                return f"[{pattern_name.upper()}]"

        # Check date patterns
        for pattern in self._date_patterns:
            try:
                datetime.strptime(value_str, pattern)
                return "[DATE]"
            except ValueError:
                continue

        # Check time patterns
        for pattern in self._time_patterns:
            try:
                datetime.strptime(value_str, pattern)
                return "[TIME]"
            except ValueError:
                continue

        return "Others"

    @staticmethod
    def _cell_to_tuple(cell_ref: str) -> Tuple[str, int]:
        """Helper function to split cell reference into column letter and row number."""
        col_str = "".join(filter(str.isalpha, cell_ref))
        row_str = "".join(filter(str.isdigit, cell_ref))
        return col_str, int(row_str) if row_str else 0

    def compress_cell_references(self, references: List[Tuple[str, str]]) -> List[str]:
        """
        Compress a list of cell references (sheet, cell) into ranges where possible.

        Args:
            references: List of tuples, where each tuple contains the sheet name and the cell reference (e.g., ("Sheet1", "A1")).

        Returns:
            List of compressed cell range strings (e.g., "Sheet1!A1:B2").
        """
        if not references:
            return []

        try:
            # Sort references by sheet, then by column and row
            sorted_refs = sorted(
                references,
                key=lambda x: (
                    x[0],  # sheet
                    column_index_from_string(
                        self._cell_to_tuple(x[1])[0]
                    ),  # column number
                    self._cell_to_tuple(x[1])[1],  # row number
                ),
            )

            ranges = []
            if not sorted_refs:
                return ranges

            current_range_start_sheet, current_range_start_cell = sorted_refs[0]
            current_range_end_sheet, current_range_end_cell = sorted_refs[0]

            def _format_range(start_sheet, start_cell, end_cell):
                if start_cell == end_cell:
                    return f"{start_sheet}!{start_cell}"
                return f"{start_sheet}!{start_cell}:{end_cell}"

            for i in range(1, len(sorted_refs)):
                sheet, cell = sorted_refs[i]
                start_col, start_row = self._cell_to_tuple(current_range_start_cell)
                end_col, end_row = self._cell_to_tuple(current_range_end_cell)
                current_col, current_row = self._cell_to_tuple(cell)

                if sheet == current_range_start_sheet:
                    if (current_col == end_col and current_row == end_row + 1) or (
                        current_row == end_row
                        and column_index_from_string(current_col)
                        == column_index_from_string(end_col) + 1
                    ):
                        # Extend current range
                        current_range_end_cell = cell
                    else:
                        # Flush current range
                        ranges.append(
                            _format_range(
                                current_range_start_sheet,
                                current_range_start_cell,
                                current_range_end_cell,
                            )
                        )
                        current_range_start_sheet, current_range_start_cell = (
                            sheet,
                            cell,
                        )
                        current_range_end_sheet, current_range_end_cell = sheet, cell
                else:
                    # Flush current range
                    ranges.append(
                        _format_range(
                            current_range_start_sheet,
                            current_range_start_cell,
                            current_range_end_cell,
                        )
                    )
                    current_range_start_sheet, current_range_start_cell = sheet, cell
                    current_range_end_sheet, current_range_end_cell = sheet, cell

            # Flush the last range
            ranges.append(
                _format_range(
                    current_range_start_sheet,
                    current_range_start_cell,
                    current_range_end_cell,
                )
            )
            return ranges

        except Exception as e:
            self.logger.error(f"Error compressing cell references: {str(e)}")
            # Return uncompressed references as fallback
            return [f"{ref[0]}!{ref[1]}" for ref in references]

    def _process_cells(
        self,
        cells: Union[
            openpyxl.worksheet.worksheet.Worksheet.rows,
            List[List[Union[str, int, float]]],
        ],
        sheet_name: str,
    ) -> Dict[str, List[str]]:
        """
        Process cells from a sheet and group them by their data type.

        Args:
            cells: Iterable of row objects (for Excel) or list of lists (for ODS/CSV).
            sheet_name: Name of the current sheet.

        Returns:
            Dictionary mapping data types to compressed cell references.
        """
        grouped_data: Dict[str, List[Tuple[str, str]]] = defaultdict(list)

        if isinstance(cells, openpyxl.worksheet.worksheet.Worksheet.rows):
            for row in cells:
                for cell in row:
                    if cell.value is not None:
                        dtype = self.recognize_data_type(cell.value)
                        if dtype == "Others":
                            grouped_data[str(cell.value)].append(
                                [sheet_name, cell.coordinate]
                            )
                        elif dtype != "Empty":
                            grouped_data[dtype].append([sheet_name, cell.coordinate])
        elif isinstance(cells, list):
            for row_idx, row in enumerate(cells, start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    if cell_value is not None:
                        dtype = self.recognize_data_type(cell_value)
                        col_letter = get_column_letter(col_idx)
                        cell_ref = f"{col_letter}{row_idx}"
                        if dtype == "Others":
                            grouped_data[str(cell_value)].append([sheet_name, cell_ref])
                        elif dtype != "Empty":
                            grouped_data[dtype].append([sheet_name, cell_ref])

        return {
            key: self.compress_cell_references(references)
            for key, references in grouped_data.items()
        }

    def parse_excel(self, file: BytesIO) -> Dict[str, List[str]]:
        """
        Parse an Excel file and group cells by their data type.

        Args:
            file: A BytesIO object containing the Excel file data.

        Returns:
            A dictionary mapping data types to compressed cell references.

        Raises:
            SpreadsheetParsingError: If there is an error parsing the Excel file.
        """
        try:
            wb = openpyxl.load_workbook(file)
            result: Dict[str, List[str]] = {}

            for sheet_name in wb.sheetnames:
                self.logger.debug(f"Processing sheet: {sheet_name}")
                ws = wb[sheet_name]
                sheet_data = self._process_cells(ws.rows, sheet_name)
                # Merge the sheet data with the overall result
                for key, value in sheet_data.items():
                    if key in result:
                        result[key].extend(value)
                    else:
                        result[key] = value

            self.logger.info("Excel parsing completed successfully")
            return result

        except openpyxl.utils.exceptions.InvalidFileException as e:
            self.logger.error(f"Error: Invalid Excel file format - {e}")
            raise SpreadsheetParsingError(f"Invalid Excel file format: {e}")
        except Exception as e:
            self.logger.error(f"Error parsing Excel file: {str(e)}")
            raise SpreadsheetParsingError(f"Error parsing Excel file: {e}")

    def parse_ods(self, file: BytesIO) -> Dict[str, List[str]]:
        """
        Parse an ODS file and group cells by their data type.

        Args:
            file: A BytesIO object containing the ODS file data.

        Returns:
            A dictionary mapping data types to compressed cell references.

        Raises:
            SpreadsheetParsingError: If there is an error parsing the ODS file.
        """
        try:
            data = pyexcel_ods.get_data(file)
            grouped_data: Dict[str, List[Tuple[str, str]]] = defaultdict(list)

            for sheet_name, sheet_data in data.items():
                self.logger.debug(f"Processing sheet: {sheet_name}")
                sheet_result = self._process_cells(sheet_data, sheet_name)
                for key, value in sheet_result.items():
                    grouped_data[key].extend(value)

            compressed_data = {
                key: self.compress_cell_references(references)
                for key, references in grouped_data.items()
            }

            self.logger.info("ODS parsing completed successfully")
            return compressed_data

        except Exception as e:
            self.logger.error(f"Error parsing ODS file: {str(e)}")
            raise SpreadsheetParsingError(f"Error parsing ODS file: {e}")

    def parse_csv(
        self, file: BytesIO, encoding: str = "utf-8", delimiter: str = ","
    ) -> Dict[str, List[str]]:
        """
        Parse a CSV file and group cells by their data type.

        Args:
            file: A BytesIO object containing the CSV file data.
            encoding: The file encoding (default: utf-8).
            delimiter: The CSV delimiter (default: comma).

        Returns:
            A dictionary mapping data types to compressed cell references.

        Raises:
            SpreadsheetParsingError: If there is an error parsing the CSV file.
        """
        try:
            grouped_data: Dict[str, List[Tuple[str, str]]] = defaultdict(list)

            text = file.getvalue().decode(encoding)
            csv_data = StringIO(text)
            reader = csv.reader(csv_data, delimiter=delimiter)

            rows = list(reader)
            sheet_result = self._process_cells(rows, "Sheet1")
            for key, value in sheet_result.items():
                grouped_data[key].extend(value)

            compressed_data = {
                key: self.compress_cell_references(references)
                for key, references in grouped_data.items()
            }

            self.logger.info("CSV parsing completed successfully")
            return compressed_data

        except Exception as e:
            self.logger.error(f"Error parsing CSV file: {str(e)}")
            raise SpreadsheetParsingError(f"Error parsing CSV file: {e}")

    def parse_file(
        self, file_url: str, verify_ssl: bool = True, **kwargs
    ) -> Dict[str, List[str]]:
        """
        Parse any supported spreadsheet file from a URL based on its extension.

        Args:
            file_url: The URL of the spreadsheet file.
            verify_ssl: Whether to verify the SSL certificate of the URL (default: True).
            **kwargs: Additional keyword arguments to pass to the specific parser function (e.g., 'encoding' and 'delimiter' for CSV).

        Returns:
            A dictionary mapping data types to compressed cell references.

        Raises:
            ValueError: If the file extension is not supported.
            requests.exceptions.RequestException: If there is an issue downloading the file.
            SpreadsheetParsingError: If there is an error parsing the file content.
        """
        try:
            self.logger.info(f"Downloading file from: {file_url}")
            response = requests.get(file_url, verify=verify_ssl, stream=True)
            response.raise_for_status()  # Raise an exception for bad status codes

            file = BytesIO(response.content)
            file_path = Path(file_url)
            extension = file_path.suffix.lower()

            self.logger.info(f"Parsing file with extension: {extension}")
            if extension in (".xlsx", ".xlsm", ".xltx", ".xltm"):
                return self.parse_excel(file)
            elif extension == ".ods":
                return self.parse_ods(file)
            elif extension == ".csv":
                return self.parse_csv(file, **kwargs)
            else:
                raise ValueError(
                    f"Unsupported file extension: {extension}. Supported extensions are: {self.SUPPORTED_EXTENSIONS}"
                )

        except requests.exceptions.RequestException as e:
            self.logger.error(f"Error downloading file from {file_url}: {e}")
            raise
        except ValueError as e:
            self.logger.error(f"Error: {e}")
            raise
        except SpreadsheetParsingError:
            raise
        except Exception as e:
            self.logger.error(
                f"An unexpected error occurred while parsing {file_url}: {e}"
            )
            raise
